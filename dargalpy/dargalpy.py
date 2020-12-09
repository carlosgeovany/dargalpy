import glob
import os
from datetime import date ,timedelta

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd




today = date.today()
date = today.today() - timedelta(days=2)
date = date.strftime("%d%b%y")

def get_xl(name,files, path):
    
    terr = []
    ccalls = []
    dtime = []
    dmiles = []
    
    for f in files:
        fo = open(f,'r', encoding='utf8')
        fl = fo.readlines()
        for l in fl:
            if "SUMMARY TOTALS:" in l:
                break
            elif "SUMMARY: CSE #" in l:
                terr.append(l[16:21])
                ccalls.append(l[62:67])
            elif "TOTAL HOURS    =" in l:
                dtime.append(l[93:99])
            elif "TRAVEL DISTANCE             =" in l:
                dmiles.append(l[61:66])
        fo.close()
    dic = dict(zip(terr, zip(terr,ccalls,dtime,dmiles)))
    df = pd.DataFrame(dic)
    df = df.T
    df = df.apply(pd.to_numeric, errors = 'coerce')
    df.columns = ['terr','ccalls','dtime','dmiles']
    df.reset_index(drop=True, inplace=True)
    xlfile = path + '\\report\\Master'+name+'.xlsm'
    xl = pd.ExcelFile(xlfile)
    xl = pd.read_excel(xl, "Time & Expenses", usecols="D")
    xl.columns = ['terr']
    xl = xl.drop(xl.index[:3])
    xl.reset_index(drop=True, inplace=True)
    xl= xl.apply(pd.to_numeric, errors = 'coerce')
    xl = xl.merge(df, on="terr", how = 'left')
    wb = load_workbook(filename=xlfile, read_only=False, keep_vba=True)
    ws = wb["Time & Expenses"]
    for index, row in xl.iterrows():
        DT = 'I%d'  % (index + 5)
        DM = 'J%d'  % (index + 5)
        CC = 'K%d'  % (index + 5)        
        ws[DT] = row[2]
        ws[DM] = row[3]
        ws[CC] = row[1]
    wb.save(path + '\\report\\' + date + name +'.xlsm')


def run_process(path):
    dameron_files = []
    munoz_files = []
    count = 0
    try:
        for filename in glob.glob(path +'\\dargal\\*.TXT'):
            if '6531' in filename:
                dameron_files.append(filename)
                count+=1
            if '6543' in filename:
                dameron_files.append(filename)
                count+=1
            if '6433' in filename:
                munoz_files.append(filename)
                count+=1
            if '6446' in filename:
                munoz_files.append(filename)
                count+=1
        
        if len(dameron_files) > 0:
            get_xl("Dameron",dameron_files, path)
        if len(munoz_files) > 0:
            get_xl("Munoz",munoz_files, path)
    
        if count == 0:
            print("No Dargal files found in "+path +'\\dargal')
        else:
            for f in glob.glob(path +'\\dargal\\*.TXT'):
                os.remove(f)
            print("Dargal files removed...")
            print("Dargal report Ready on " + path + '\\report')
    except Exception as e:
        print (f"Exception ocurred: {e}")
